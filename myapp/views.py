from django.shortcuts import render,HttpResponse,get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from .models import *
from rest_framework.parsers import JSONParser
from django.middleware import csrf
from django.contrib.auth import authenticate
from django.core.exceptions import ObjectDoesNotExist
from .helper import send_forget_password_mail,send_permission_mail
import json
import random
from datetime import datetime,timedelta


@csrf_exempt
def signup_email(request):
    if request.method == 'POST':
        jsondata = JSONParser().parse(request)
        email = jsondata.get('Email')
        try:
            xx = Myuser.objects.filter(email=email)
            if xx.exists():
                return JsonResponse({'message':'Email id already in used'}, status=406)
            else:
                user = Myuser.objects.create(email=email)
                otp_code = generate_otp()
                print(otp_code)
                user.otp_code = otp_code
                send_forget_password_mail(email, otp_code) 
                user.save()
                return JsonResponse({'message':'Email send successfully'})
            
        except:
            return JsonResponse({'message':'Error'})
def generate_otp():
    return str(random.randint(100000, 999999))

@csrf_exempt
def signup(request):
    if request.method == 'POST':
        try:
            jsondata = JSONParser().parse(request)
            age = jsondata.get('age')
            privacy = jsondata.get('privacy')
            receive_news = jsondata.get('receive_news')
            email = jsondata.get('Email')
            nick_name = jsondata.get('nick_name')
            
            
            print(age,privacy,receive_news,email,nick_name)
            csrf_token = csrf.get_token(request)
            user, created = Myuser.objects.get_or_create(email=email)
            user.age = age
            user.privacy=privacy
            user.receive_news=receive_news
            user.nick_name=nick_name
            user.csrf_token=csrf_token
            
            # user.save()
            print(csrf_token)
            
            return JsonResponse({'message': 'Data saved successfully',
                                 'mail':email,'csrf_token':csrf_token,
                                 "image":user.image.url if user.image else None,
                                    'nick_name':user.nick_name,
                                    'permission':['Admin']
                                    })
        except :
            return JsonResponse({'message': 'The Email address is in used'})
    else:
        return JsonResponse({'message': 'Invalid request method'})

@csrf_exempt
def photoupload(request):
    if request.method == 'POST':
        email = request.POST.get('Email')
        photo = request.FILES.get('photo')
        
        print(email)
        print(photo)
        if not email:
            return JsonResponse({'message': 'Email is required'}, status=400)
        
        if not photo:
            return JsonResponse({'message': 'Photo is required'}, status=400)
        
        try:
            user = Myuser.objects.filter(email=email).first()
            if user:
                user.image = photo
                user.save()
                return JsonResponse({'message': 'Photo updated successfully'})
            else:
                return JsonResponse({'message': 'User not found'}, status=404)
        
        except Exception as e:
            return JsonResponse({'message': 'An error occurred', 'error': str(e)}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)

        
        
        
@csrf_exempt
def signup_otp_verify(request):
    if request.method == 'POST':
        jsondata = JSONParser().parse(request)
        otp = jsondata.get('otp')
        try:
            user = Myuser.objects.filter(otp_code=otp)
            if user.exists():
                return JsonResponse({'message':'login data successfull','permission':'Admin'})
            else:
                return JsonResponse({'message':'Mail id doesnot exists'})
        except:
            return JsonResponse({'message':'error'})


@csrf_exempt
def login(request):
    if request.method == 'POST':
        jsondata = JSONParser().parse(request)
        email = jsondata.get('Email')
        try:
            user = Myuser.objects.get(email=email)
        except Myuser.DoesNotExist:
            try:
                user = Permiteduser.objects.get(email=email)
            except Permiteduser.DoesNotExist:
                 return JsonResponse({'message': 'Email not found'}, status=404) 
            print(user)
        otp_code = generate_otp()
        print(otp_code)
        user.otp_code = otp_code
        user.save()
        send_forget_password_mail(email, otp_code)
        # return JsonResponse({'message': 'Email sent','csrf_token':user.csrf_token}, status=200)
        return JsonResponse({'message': 'Email sent'}, status=200)    
    
  
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=400) 

def generate_otp():
    return str(random.randint(100000, 999999))

@csrf_exempt
def login_otp_verify(request):
    if request.method == 'POST':
        jsondata = JSONParser().parse(request)
        otp = jsondata.get('otp')
        try: 
            user = Myuser.objects.filter(otp_code=otp).first()
            if user:
                return JsonResponse({
                    'message': 'login with super user successful',
                    'csrf_token': user.csrf_token,
                    'mail': user.email,
                    'image': user.image.url if user.image else None,
                    'nick_name': user.nick_name,
                    'permission':['Admin']
                })

            user = Permiteduser.objects.filter(otp_code=otp).first()
            if user:
                try:
                    role = Role.objects.get(role=user.myteam)  
                except Role.DoesNotExist:
                    return JsonResponse({'message': 'Role does not exist for this user'}, status=404)
                
                return JsonResponse({
                    'message': 'login with permitted user successful',
                    'mail': user.email,
                    'image': user.image.url if user.image else None,
                    'permission':role.permissions
                })

            return JsonResponse({'message': 'OTP does not exist'}, status=401)
        
        except Exception as e:
            return JsonResponse({'message': 'error', 'details': str(e)}, status=500)

    else:
        return JsonResponse({'message': 'Invalid request method'}, status=400)

########################################## TEAM ##############################################################

@csrf_exempt
def create_team(request):
    if request.method == 'POST':
        jsondata = JSONParser().parse(request)
        email = jsondata.get('Email')
        team_name = jsondata.get('team_name')
        team_mode = jsondata.get('team_mode')
        # print(email)
        # print(team_name)
        # print(team_mode)
        try:
            result = Myuser.objects.get(email=email)
            print(result)
            if result:
                obj = Myteam.objects.create(name=team_name,team_mode=team_mode,email=result)
                obj.save()
                return JsonResponse({'message':'Team create successfull....'})
            else:
                return JsonResponse({'message':'Email not found..'})
        except:
            return JsonResponse({'message':'An error occured..'})
    else:
        return JsonResponse({'message':'error.'})
    
@csrf_exempt
def company_information(request):
    if request.method == 'POST':
        jsondata = JSONParser().parse(request)
        company_name = jsondata.get('company_name')
        business_regd_number = jsondata.get('business_regd_number')
        address = jsondata.get('address')
        company_representative = jsondata.get('company_representative')
        phone = jsondata.get('phone')
        team_id = jsondata.get('team_id')
        
        print(company_name,business_regd_number,address,company_representative,phone,team_id)
        try:
            team = Myteam.objects.get(id=team_id)
            if team:
                CompanyInformation.objects.create(
                    company_name = company_name,
                    business_regd_number = business_regd_number,
                    address = address,
                    company_representative =company_representative,
                    phone = phone,
                    team=team
                )
                return JsonResponse({'message':'Data save successfull'})
        except Myteam.DoesNotExist:
            return JsonResponse({'message': 'Team not found'}, status=404)
        except Exception as e:
            return JsonResponse({'message': f'Error: {str(e)}'}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)
    
@csrf_exempt
def delete_team(request, id):
    if request.method == 'DELETE':
        try:
            result = Myteam.objects.filter(id=id)
            if result.exists():
                result.delete()
                return JsonResponse({'message': 'Delete successful'})
            else:
                return JsonResponse({'message': 'ItemAttribute not found'}, status=404)
        except Exception as e:
            return JsonResponse({'message': str(e)}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)
        
      
@csrf_exempt
def team_list(request, Email):
    if request.method == 'GET':
        try:
            user = Myuser.objects.get(email=Email)
            result = Myteam.objects.filter(email=user)
            if result.exists():
                data = []
                for team in result:
                    item_count = Item_Information.objects.filter(myteam=team).count()
                    value = {
                        'id': team.id,
                        'team_name': team.name,
                        'team_mode': team.team_mode,
                        'team_note': team.team_note,
                        'image': team.image.url if team.image else None,
                        'timezone': team.timezone,
                        'item_count': item_count
                    }
                    data.append(value)
                return JsonResponse(data, safe=False)
            else:
                return JsonResponse({'message': 'No teams found for this user'}, status=404)
        except Myuser.DoesNotExist:
            try:
                permiteduser = Permiteduser.objects.filter(email=Email)
                if permiteduser.exists():
                    data = []
                    for p_user in permiteduser:
                        team = p_user.myteam
                        item_count = Item_Information.objects.filter(myteam=team).count()
                        value = {
                            'id': team.id,
                            'team_name': team.name,
                            'team_mode': team.team_mode,
                            'team_note': team.team_note,
                            'image': team.image.url if team.image else None,
                            'timezone': team.timezone,
                            'item_count': item_count
                        }
                        data.append(value)
                    return JsonResponse(data, safe=False)
                else:
                    return JsonResponse({'message': 'No permitted teams found for this email'}, status=404)
            except Permiteduser.DoesNotExist:
                return JsonResponse({'message': 'Email not found in Myuser or Permiteduser tables'}, status=404)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)
    

@csrf_exempt
def update_team(request, id):
    if request.method == 'POST':
        team_note = request.POST.get('team_note')
        image = request.FILES.get('image')
        timezone = request.POST.get('timezone')
        name = request.POST.get('name')
        company_name = request.POST.get('companyName')
        registration = request.POST.get('registration')
        address = request.POST.get('address')
        representative = request.POST.get('representative')
        phone = request.POST.get('phone')
        currency = request.POST.get('currency')
        items = request.POST.get('items')

        try:
            team = Myteam.objects.get(id=id)
            if team_note:
                team.team_note = team_note
            if image:
                team.image = image
            if timezone:
                team.timezone = timezone
            if name:
                team.name = name
            if company_name:
                team.company_name = company_name
            if registration:
                team.registration = registration
            if address:
                team.address = address
            if representative:
                team.representative = representative
            if phone:
                team.phone = phone
            if currency:
                team.currency = currency
            if items:
                team.items = items

            team.save()

            return JsonResponse({'message': 'Team information has been updated!'})
        except Myteam.DoesNotExist:
            return JsonResponse({'message': 'Team not found'}, status=404)
        except Exception as e:
            return JsonResponse({'message': f'Error: {str(e)}'}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)
    
@csrf_exempt  
def get_team_details(request, id):
    if request.method == 'GET':
        try:
            team = Myteam.objects.get(id=id)
            team_data = {
                'id': team.id,
                'name': team.name,
                'team_note': team.team_note,
                'timezone': team.timezone,
                'company_name': team.company_name,
                'registration': team.registration,
                'address': team.address,
                'representative': team.representative,
                'phone': team.phone,
                'currency': team.currency,
                'items': team.items,
                'team_mode': team.team_mode,
                'image': team.image.url,
                
            }
            return JsonResponse(team_data)
        except Myteam.DoesNotExist:
            return JsonResponse({'message': 'Team not found'}, status=404)
        except Exception as e:
            return JsonResponse({'message': f'Error: {str(e)}'})
   


################################################################# item list ###################################################################################
        
@csrf_exempt
def add_item(request, id):
    if request.method == 'POST':
        try:
            Name = request.POST.get('Name')
            Barcode = request.POST.get('Barcode')
            cost = request.POST.get('cost')
            Price = request.POST.get('price')
            image = request.FILES.get('image')
            initial_quantity = request.POST.get('initial_quantity')

            print(id, Name, Barcode, cost, Price, image, initial_quantity)

            user = Myteam.objects.get(id=id)
            
            data = Item_Information.objects.create(
                Name=Name,
                Barcode=Barcode,
                cost=cost,
                Price=Price,
                image=image,
                initial_quantity=initial_quantity,
                myteam=user
            )
            
            attributes = request.POST.get('attributes', '[]')
            attributes = json.loads(attributes) 
            print(attributes)
            
            for attr in attributes:
                attribute_id = attr.get('attribute_id')
                attribute_name = attr.get('attribute_name')
                attribute_value = attr.get('attribute_value')
                
                if attribute_id:
                    ItemAttribute.objects.filter(id=attribute_id).update(
                        attribute_name=attribute_name,
                        attribute_value=attribute_value,
                        name=user,
                        item=data
                    )
                else:
                    ItemAttribute.objects.create(
                        attribute_name=attribute_name,
                        attribute_value=attribute_value,
                        name=user,
                        item=data
                    )
                
            response_data = {
                "Name": data.Name,
                "Barcode": data.Barcode,
                "cost": data.cost,
                "Price": data.Price,
                "image": data.image.url if data.image else None,
                "initial_quantity": data.initial_quantity,
                "attributes": attributes
            }
            
            return JsonResponse(response_data, status=200)
        
        except Myteam.DoesNotExist:
            return JsonResponse({'message': 'User not found'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'message': 'Invalid JSON for attribute field'}, status=400)
        except Exception as e:
            return JsonResponse({'message': 'error', 'details': str(e)}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Item_Information, ItemAttribute
import json

@csrf_exempt
def edit_item(request, id):  # Attribute edit successful, add needs to focus
    if request.method == 'POST':
        Name = request.POST.get('Name', None)
        Barcode = request.POST.get('Barcode', None)
        cost = request.POST.get('cost', None)
        Price = request.POST.get('Price', None)
        image = request.FILES.get('image', None)
        initial_quantity = request.POST.get('initial_quantity', None)
        
        try:
            data = Item_Information.objects.get(id=id)
            
            if Name is not None:
                data.Name = Name
            if Barcode is not None:
                data.Barcode = Barcode
            if cost is not None:
                data.cost = cost
            if Price is not None:
                data.Price = Price
            if image is not None:
                data.image = image
            if initial_quantity is not None:
                data.initial_quantity = initial_quantity
            
            data.save()
            print("ok")
            
            attributes = request.POST.get('attributes', '[]')
            attributes = json.loads(attributes)
            for attr in attributes:
                attribute_id = attr.get('attribute_id', None)
                attribute_name = attr.get('attribute_name', None)
                attribute_value = attr.get('attribute_value', None)
                
                
                if attribute_id is not None:
                    # Update existing attribute
                    ItemAttribute.objects.filter(id=attribute_id, item=data).update(
                        attribute_value=attribute_value,
                    )
                    print("ok done")
                else:
                    # Create new attribute
                    ItemAttribute.objects.create(
                        attribute_name=attribute_name,
                        attribute_value=attribute_value,
                        name=data.myteam,
                        item=data
                    )
                    print("data save")
                
            return JsonResponse({'message': 'Data saved successfully'}, status=200)
        
        except Item_Information.DoesNotExist:
            return JsonResponse({'message': 'Item not found'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'message': 'Invalid JSON for attribute field'}, status=400)
        except Exception as e:
            return JsonResponse({'message': 'Error', 'details': str(e)}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)

    
@csrf_exempt
def item(request,id):
    if request.method == 'POST':
        attribute_value = request.POST.get('attribute_value')
        attribute_name = request.POST.get('attribute_name')  
        try:
            item = Item_Information.objects.get(id=id)
            myteam = item.myteam  
            
            ItemAttribute.objects.create(
                attribute_name=attribute_name,
                attribute_value=attribute_value,
                name=myteam,
                item=item
            )
            
            return JsonResponse({'message': 'Attribute created successfully'}, status=200)
        except Item_Information.DoesNotExist:
            return JsonResponse({'message': 'Item not found'}, status=404)
        except Exception as e:
            return JsonResponse({'message': 'Error', 'details': str(e)}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)
        

@csrf_exempt
def delete_item(request,id):
    if request.method == 'DELETE':
        try:
            result = Item_Information.objects.filter(id=id)
            if result.exists():
                result.delete()
                return JsonResponse({'message': 'Delete successful'},status=200)
            else:
                return JsonResponse({'message': 'ItemAttribute not found'}, status=404)
        except Exception as e:
            return JsonResponse({'message': str(e)}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)
    
@csrf_exempt
def transaction_item(request, id):
    if request.method == 'GET':
        try:
            data = Item_Information.objects.get(id=id)
            stockin = Stockin.objects.filter(myteam=data)
            stockout = Stockout.objects.filter(myteam=data)
            adjust = Adjust.objects.filter(myteam=data)
            transactions = []
            
            for i in stockin:
                transaction_in = {
                    "type": "Stock In",
                    "supplier":i.supplier,
                    "initial_quantity": i.initial_quantity,
                    "available": data.initial_quantity,
                    "created_at": i.created_at
                }
                transactions.append(transaction_in)
                
            for i in stockout:
                transaction_out = {
                    "type": "Stock Out",
                    "customer":i.customer,
                    "initial_quantity": i.initial_quantity,
                    "available": data.initial_quantity,
                    "created_at": i.created_at
                }
                transactions.append(transaction_out)
            for i in adjust:
                ad = {
                    "type": "Adjust",
                    "initial_quantity": i.initial_quantity,
                    "available": data.initial_quantity,
                    "created_at": i.created_at
                }
                transactions.append(ad)
            
                
            return JsonResponse(transactions, safe=False)
        except Item_Information.DoesNotExist:
            return JsonResponse({'message': 'Item not found'}, status=404)
        except Exception as e:
            return JsonResponse({'message': 'error', 'details': str(e)}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)

     
@csrf_exempt
def initial_quantity(request,id):
    if request.method == 'GET':
        try:
            item = Item_Information.objects.get(id=id)
            return JsonResponse({'initial_quantity':item.initial_quantity}, safe=False)
        except Item_Information.DoesNotExist:
            return JsonResponse({'message':'Item not found'}, status=404)       
        except Exception as e:
            return JsonResponse({'message':'An error occured', 'details': str(e)}, status=500)
    else:
        return JsonResponse({'error':'Invalid request method'}, status=405)

    
############################## Attribute ############################################################################
@csrf_exempt
def add_attribute(request, id):                                 
    if request.method == 'POST':
        jsondata = JSONParser().parse(request)
        attribute_name = jsondata.get("attribute_name")
        attribute_type = jsondata.get("attribute_type")
        try:
            xx = Myteam.objects.get(id=id)
            
            ItemAttribute.objects.create(
                attribute_name=attribute_name,
                attribute_type=attribute_type,
                name=xx)
            return JsonResponse({'message':'data save successfull'})
        except:
            return JsonResponse({'message':'error'})
    else:
        return JsonResponse({'message':'error'})
@csrf_exempt
def send_attribute(request,id):
    if request.method == 'GET':
        
        try:
            xx = Myteam.objects.get(id=id)
            attributes = ItemAttribute.objects.filter(name=xx)
            list = []
            for i in attributes:
                data = {
                    "id":i.id,
                    "attribute_name":i.attribute_name,
                    "attribute_type":i.attribute_type
                }
                list.append(data)
            return JsonResponse(list, safe=False)
        except:
            return JsonResponse({'message':'error'})
        
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from rest_framework.parsers import JSONParser

@csrf_exempt
def edit_attribute(request, id):
    if request.method == 'POST':
        jsondata = JSONParser().parse(request)
        attribute_name = jsondata.get("attribute_name")

        try:
            myteam = ItemAttribute.objects.get(id=id)
           
            items = ItemAttribute.objects.filter(id=id)
            for item in items:
                item.attribute_name = attribute_name
                item.save()

            return JsonResponse({'message': 'Data updated successfully'})

        except Myteam.DoesNotExist:
            return JsonResponse({'message': 'Myteam not found'}, status=404)

        except Exception as e:
            return JsonResponse({'message': str(e)}, status=500)

    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)
    
@csrf_exempt
def delete_attribute(request, id):
    if request.method == 'DELETE':
        try:
            result = ItemAttribute.objects.filter(id=id)
            if result.exists():
                result.delete()
                return JsonResponse({'message': 'Delete successful'})
            else:
                return JsonResponse({'message': 'ItemAttribute not found'}, status=404)
        except Exception as e:
            return JsonResponse({'message': str(e)}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)

        
        
@csrf_exempt
def item_list_view(request, id):
    if request.method == 'GET':
        try:
            result = Myteam.objects.get(id=id)
            items = Item_Information.objects.filter(myteam=result)
            data = []

            for item in items:
                attributes = ItemAttribute.objects.filter(item=item)
                attribute_data = [
                    {"attribute_name": attr.attribute_name, "attribute_value": attr.attribute_value}
                    for attr in attributes
                ]
                
                stockins = Stockin.objects.filter(myteam=item)
                rack_nos = list(set(i.rack_no for i in stockins))
                memos = [stockin.memo for stockin in stockins]                                 #added new

                item_data = {
                    "id": item.id,
                    "Name": item.Name,
                    "Barcode": item.Barcode,
                    "cost": item.cost,
                    "Price": item.Price,
                    "memo": memos,
                    "initial_quantity": item.initial_quantity,
                    "image": item.image.url if item.image else None,
                    "attributes": attribute_data,
                    "rack_no":rack_nos,
                    "created_at":item.created_at
                }

                data.append(item_data)

            return JsonResponse(data, safe=False)
        
        except Myteam.DoesNotExist:
            return JsonResponse({'message': 'Team not found'}, status=404)
        except Exception as e:
            return JsonResponse({'message': str(e)}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)
        


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q
from rest_framework.parsers import JSONParser
from .models import Item_Information

@csrf_exempt
def groupby(request):
    if request.method == 'POST':
        try:
            jsondata = JSONParser().parse(request)
            Name = jsondata.get("Name", None)
            Barcode = jsondata.get("Barcode", None)
            cost = jsondata.get("cost", None)
            
            # Construct the query based on the provided key
            query = Q()
            if Name is not None:
                query = Q(Name=Name)
            elif Barcode is not None:
                query = Q(Barcode=Barcode)
            elif cost is not None:
                query = Q(cost=cost)
            else:
                return JsonResponse({'message': 'No valid filter key provided'}, status=400)
            
            items = Item_Information.objects.filter(query)
            data = [] 
            for item in items:
                item_data = {
                    "Name": item.Name,
                    "Barcode": item.Barcode,
                    "cost": item.cost,
                    "Price": item.Price,
                    "image": item.image.url if item.image else None
                }
                data.append(item_data)

            return JsonResponse(data, safe=False)
        except Exception as e:
            return JsonResponse({'message': str(e)}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)

    # else:
    #     return JsonResponse({'message': 'Invalid request method'}, status=405)
        
# class ExcelUploadView(APIView):
#     def post(self, request):
#         if 'file' not in request.FILES:
#             return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
#         file = request.FILES['file']
#         if not file.name.endswith('.xlsx'):
#             return Response({'error': 'Invalid file type'}, status=status.HTTP_400_BAD_REQUEST)
        
#         fs = FileSystemStorage()
#         filename = fs.save(file.name, file)
#         file_path = fs.path(filename)
        
#         wb = load_workbook(file_path)
#         ws = wb.active

#         data = []
#         for row in ws.iter_rows(min_row=2, values_only=True):
#             obj = GFG(
#                 name=row[0],
#                 contact=row[1],
#                 address=row[2]
#             )
#             data.append(obj)

#         GFG.objects.bulk_create(data)
#         return Response({'message': 'File imported successfully'}, status=status.HTTP_201_CREATED)


from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

from .models import Item_Information, ItemAttribute, Myteam

@csrf_exempt
def add_item_excel_upload(request,id):
    if request.method == 'POST':
        try:
            myteam = Myteam.objects.get(id=id)
        except Myteam.DoesNotExist:
            return JsonResponse({'error': 'Myteam with ID does not exist'}, status=400)

        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file provided'}, status=400)

        file = request.FILES['file']

        if not file.name.endswith('.xlsx'):
            return JsonResponse({'error': 'Invalid file type'}, status=400)

        fs = FileSystemStorage()
        filename = fs.save(file.name, file)
        file_path = fs.path(filename)

        wb = load_workbook(file_path)
        ws = wb.active

        items_data = []
        attributes_data = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                name = row[0]
                barcode = row[1]
                cost = row[2]
                price = row[3]
                initial_quantity = row[4]
                attribute_name = row[5]
                attribute_value = row[6]

                item_info = Item_Information(
                    Name=name,
                    Barcode=barcode,
                    cost=cost,
                    Price=price,
                    initial_quantity=initial_quantity,
                    myteam=myteam
                )
                item_info.save()
                items_data.append(item_info)

                item_attribute = ItemAttribute(
                    attribute_name=attribute_name,
                    attribute_value=attribute_value,
                    name=myteam,
                    item=item_info  
                )
                attributes_data.append(item_attribute)

            except Exception as e:
                return JsonResponse({'error': str(e)}, status=500)

        ItemAttribute.objects.bulk_create(attributes_data)

        return JsonResponse({'message': 'File imported successfully'}, status=201)



    
######################################################### Partner ######################################################################
        
@csrf_exempt
def add_partner(request, id):
    if request.method == 'POST':
        try:
            jsondata = JSONParser().parse(request)
            type = jsondata.get('type')
            name = jsondata.get('name')
            mobile = jsondata.get('mobile')
            email = jsondata.get('email')
            address = jsondata.get('address')
            memo = jsondata.get('memo')
            favorite = jsondata.get('favorite')

            myteam = Myteam.objects.get(id=id)

            if type == 'Supplier':
                result = Supplier.objects.create(
                    type=type,
                    name=name,
                    mobile=mobile,
                    email=email,
                    address=address,
                    memo=memo,
                    favorite=favorite,
                    foreign=myteam  
                )
                return JsonResponse({'message': 'Supplier data saved successfully'})
            elif type == 'Customer':
                result = Customer.objects.create(
                    type=type,
                    name=name,
                    mobile=mobile,
                    email=email,
                    address=address,
                    memo=memo,
                    favorite=favorite,
                    foreign=myteam  
                )
                return JsonResponse({'message': 'Customer data saved successfully'})
            else:
                return JsonResponse({'message': 'Invalid type provided'}, status=400)
        except Myteam.DoesNotExist:
            return JsonResponse({'message': 'Myteam with given ID does not exist'}, status=400)
        except Exception as e:
            return JsonResponse({'message': str(e)}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)
    
@csrf_exempt
def partner_view(request, id):
    if request.method == 'GET':
        try:
            myteam = Myteam.objects.get(id=id)
            
            suppliers = Supplier.objects.filter(foreign=myteam)
            customers = Customer.objects.filter(foreign=myteam)

            # Prepare the list of suppliers
            supplier_list = []
            for supplier in suppliers:
                supplier_data = {
                    "id": supplier.id,
                    "type": supplier.type,
                    "name": supplier.name,
                    "mobile": supplier.mobile,
                    "email": supplier.email,
                    "address": supplier.address,
                    "memo": supplier.memo,
                    "favorite": supplier.favorite
                }
                supplier_list.append(supplier_data)

            # Prepare the list of customers, including supplier list
            customer_list = []
            for customer in customers:
                customer_data = {
                    "id": customer.id,
                    "type": customer.type,
                    "name": customer.name,
                    "mobile": customer.mobile,
                    "email": customer.email,
                    "address": customer.address,
                    "memo": customer.memo,
                    "favorite": customer.favorite,
                    
                }
                customer_list.append(customer_data)
                
            response_data = {
                "suppliers":supplier_list,
                "customers":customer_list
            }

            return JsonResponse(response_data, safe=False)
        
        except Myteam.DoesNotExist:
            return JsonResponse({'message': 'Myteam with given ID does not exist'}, status=404)
        except Exception as e:
            return JsonResponse({'message': str(e)}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)
    
@csrf_exempt
def partner_edit(request, id):
    if request.method == 'POST':
        jsondata = JSONParser().parse(request)
        type = jsondata.get('type')
        name = jsondata.get('name', None)
        mobile = jsondata.get('mobile', None)
        email = jsondata.get('email',None)
        address = jsondata.get('address',None)
        memo = jsondata.get('memo',None)

        try:
            if type == 'Supplier':
                supplier = Supplier.objects.get(id=id)

                if name is not None:
                    supplier.name = name
                if mobile is not None:
                    supplier.mobile = mobile 
                if email is not None:
                    supplier.email = email
                if address is not None:
                    supplier.address = address
                if memo is not None:
                    supplier.memo = memo
                
                supplier.save()
                
            elif type == 'Customer':
                supplier = Customer.objects.get(id=id)

                if name is not None:
                    supplier.name = name
                if mobile is not None:
                    supplier.mobile = mobile 
                if email is not None:
                    supplier.email = email
                if address is not None:
                    supplier.address = address
                if memo is not None:
                    supplier.memo = memo
                
                supplier.save()
            else:
                return JsonResponse({'message':'error'})

            return JsonResponse({'message': 'Data updated successfully'})

        except Supplier.DoesNotExist:
            return JsonResponse({'message': 'Supplier not found'}, status=404)

        except Exception as e:
            return JsonResponse({'message': str(e)}, status=500)

    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)
            
            
            
@csrf_exempt
def partner_delete(request, id):
    if request.method == 'DELETE':
        try:
            jsondata = JSONParser().parse(request)
            type = jsondata.get('type')

            print(type,id)
            if type == 'Customer':
                customer_exists = Customer.objects.filter(id=id).exists()
                if customer_exists:
                    Customer.objects.filter(id=id).delete()
                    return JsonResponse({'message': 'Customer deleted successfully'})
            elif type == 'Supplier':
                supplier_exists = Supplier.objects.filter(id=id).exists()
                if supplier_exists:
                    Supplier.objects.filter(id=id).delete()
                    return JsonResponse({'message': 'Supplier deleted successfully'})

            return JsonResponse({'message': 'Supplier or Customer not found'}, status=404)

        except Exception as e:
            return JsonResponse({'message': str(e)}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)
       
#################################################################### Stock In ###################################################################################

@csrf_exempt
def stock_in(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            supplier = data.get('supplier')
            created_at = data.get('created_at')
            memo = data.get('memo')
            items = data.get('items')
            
            # created_at = datetime.strptime(created_at_str, "%B %d, %Y, %I:%M %p")
            transaction = Transaction.objects.create()
            
            print(supplier,created_at,memo,items)
            for item in items:
                id = item.get('id')
                initial_quantity = item.get('initial_quantity')
                # present_quantity = item.get('present_quantity')                         # newly added
                rack_no = item.get('rack_no')

                user = Item_Information.objects.get(id=id)
                
                Stockin.objects.create(
                    supplier=supplier,
                    created_at=created_at,
                    initial_quantity=initial_quantity,
                    present_quantity = initial_quantity,
                    memo=memo,
                    rack_no = rack_no,
                    myteam=user,
                    transaction=transaction
                )   
                
                user.initial_quantity += int(initial_quantity)
                user.save()
            
            return JsonResponse({'message': 'Data saved successfully for all items.'})
        except Item_Information.DoesNotExist:
            return JsonResponse({'message': 'User not found'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'message': 'Invalid JSON for attribute field'}, status=400)
        except Exception as e:
            return JsonResponse({'message': 'error', 'details': str(e)}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)
    
    
@csrf_exempt
def rack_stockin_view(request, id):                                       #newly modified
    if request.method == 'GET':
        try:
            item = Item_Information.objects.get(id=id)
            stockin_records = Stockin.objects.filter(myteam=item)
            data = []
            seen_racks = set()
            for stockin in stockin_records:
                if stockin.present_quantity == 0:
                    stockin.rack_no = None
                    stockin.save()
                elif stockin.rack_no and stockin.rack_no not in seen_racks:
                    seen_racks.add(stockin.rack_no)
                    data.append({
                        "rack": stockin.rack_no
                    })
            return JsonResponse(data, safe=False)
        except Item_Information.DoesNotExist:
            return JsonResponse({'message': 'Item not found'}, status=404)
        except Exception as e:
            return JsonResponse({'message': 'error', 'details': str(e)}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)
    
@csrf_exempt
def stockin_excel_upload(request, id):
    if request.method == 'POST':
        if 'file' not in request.FILES:
            return JsonResponse({'message': 'File is required.'}, status=400)

        file = request.FILES['file']

        if not file.name.endswith('.xlsx'):
            return JsonResponse({'message': 'Invalid file type'}, status=400)

        fs = FileSystemStorage()
        filename = fs.save(file.name, file)
        file_path = fs.path(filename)

        wb = load_workbook(file_path)
        ws = wb.active

        try:
            for row in ws.iter_rows(min_row=2, values_only=True):
                supplier = row[0]
                name = row[1]
                initial_quantity = row[2]
                created_at = row[3]
                memo = row[4] if len(row) > 4 else ""
                rack_no = row[5] if len(row) > 5 else None
                
                # Automatically use the last created transaction if transaction_id is not provided
                transaction = Transaction.objects.last()
                
                items = Item_Information.objects.filter(Name=name)

                if not items.exists():
                    return JsonResponse({'message': f'Item with name {name} not found'}, status=404)

                for item in items:
                    Stockin.objects.create(
                        supplier=supplier,
                        created_at=created_at,
                        initial_quantity=initial_quantity,
                        present_quantity=initial_quantity,  # Set present_quantity equal to initial_quantity
                        memo=memo,
                        rack_no=rack_no,
                        myteam=item,
                        transaction=transaction
                    )

                    item.initial_quantity += int(initial_quantity)
                    item.save()

            return JsonResponse({'message': 'Data saved successfully for all items.'}, status=201)

        except Exception as e:
            return JsonResponse({'message': 'error', 'details': str(e)}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)


############################################################ Stock out #######################################################################################

@csrf_exempt
def stock_out(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            customer = data.get('customer')
            created_at = data.get('created_at')
            memo = data.get('memo')
            items = data.get('items')
            
            # created_at = datetime.strptime(created_at_str, "%B %d, %Y, %I:%M %p")
            transaction = Transaction.objects.create()
            
            print(customer,created_at,memo,items)
            for item in items:
                id = item.get('id')
                initial_quantity = item.get('initial_quantity')
                rack_no = item.get('rack_no')

                user = Item_Information.objects.get(id=id)
                
                Stockout.objects.create(
                    customer=customer,
                    created_at=created_at,
                    initial_quantity=-int(initial_quantity),
                    memo=memo,
                    rack_no = rack_no,
                    myteam=user,
                    transaction=transaction
                )   
                
                user.initial_quantity -= int(initial_quantity)
                user.save()
                
                stockin_items = Stockin.objects.filter(myteam=user, rack_no=rack_no)            #newly added
                for stockin in stockin_items:
                    if stockin.present_quantity >= int(initial_quantity):
                        stockin.present_quantity -= int(initial_quantity)
                        stockin.save()
                        break
                    else:
                        initial_quantity -= stockin.present_quantity
                        stockin.present_quantity = 0
                        stockin.save()
            
            return JsonResponse({'message': 'Data saved successfully for all items.'})
        except Item_Information.DoesNotExist:
            return JsonResponse({'message': 'User not found'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'message': 'Invalid JSON for attribute field'}, status=400)
        except Exception as e:
            return JsonResponse({'message': 'error', 'details': str(e)}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)
    
    
@csrf_exempt
def rack_stockout_view(request, id):
    if request.method == 'GET':
        try:
            item = Item_Information.objects.get(id=id)
            stockin_records = Stockin.objects.filter(myteam=item)
            data = []
            seen_racks = set()
            for stockin in stockin_records:
                if stockin.rack_no not in seen_racks:
                    seen_racks.add(stockin.rack_no)
                    data.append({
                        "rack": stockin.rack_no
                    })
            return JsonResponse(data, safe=False)
        except Item_Information.DoesNotExist:
            return JsonResponse({'message': 'Item not found'}, status=404)
        except Exception as e:
            return JsonResponse({'message': 'error', 'details': str(e)}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)
    
    
@csrf_exempt
def stockout_excel_upload(request,id):
    if request.method == 'POST':
        if 'file' not in request.FILES:
            return JsonResponse({'message': 'File is required.'}, status=400)

        file = request.FILES['file']

        if not file.name.endswith('.xlsx'):
            return JsonResponse({'message': 'Invalid file type'}, status=400)

        fs = FileSystemStorage()
        filename = fs.save(file.name, file)
        file_path = fs.path(filename)

        wb = load_workbook(file_path)
        ws = wb.active

        try:
            for row in ws.iter_rows(min_row=2, values_only=True):
                customer = row[0]
                name = row[1]
                initial_quantity = row[2]
                created_at = row[3]
                memo = row[4] if len(row) > 4 else ""

                # Ensure all required data is present
                if not customer or not name or initial_quantity is None or not created_at:
                    return JsonResponse({'message': 'Invalid data in file.'}, status=400)

                try:
                    items = Item_Information.objects.filter(Name=name)
                    print(items)
                    if not items.exists():
                        return JsonResponse({'message': f'Item with name {name} not found'}, status=404)

                    for item in items:
                        Stockout.objects.create(
                            customer=customer,
                            created_at=created_at,
                            initial_quantity=initial_quantity,
                            memo=memo,
                            myteam=item  
                        )

                        item.initial_quantity -= int(initial_quantity)
                        item.save()

                except Exception as e:
                    return JsonResponse({'message': 'error', 'details': str(e)}, status=500)

            return JsonResponse({'message': 'Data saved successfully for all items.'}, status=201)

        except Exception as e:
            return JsonResponse({'message': 'error', 'details': str(e)}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)
    
############################################################# Adjust #####################################################################################

@csrf_exempt
def adjust(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            memo = data.get('memo')
            items = data.get('items')
            print(memo, items)
            transaction = Transaction.objects.create()
            
            for item in items:
                id = item.get('id')
                initial_quantity = item.get('initial_quantity')
                rack_no = item.get('rack_no')

                user = Item_Information.objects.get(id=id)
                
                Adjust.objects.create(
                    initial_quantity=initial_quantity,
                    memo=memo,
                    myteam=user,
                    transaction=transaction,
                    rack_no = rack_no
                )   
                
                user.initial_quantity = int(initial_quantity)
                user.save()
            
            return JsonResponse({'message': 'Data saved successfully for all items.'})
        except Item_Information.DoesNotExist:
            return JsonResponse({'message': 'User not found'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'message': 'Invalid JSON for attribute field'}, status=400)
        except Exception as e:
            return JsonResponse({'message': 'error', 'details': str(e)}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)
    


################################################# Summry ####################################################################
# @csrf_exempt
# def summary(request, id):
#     if request.method == 'GET':
#         try:
#             result = Myteam.objects.get(id=id)
#             item_infos = Item_Information.objects.filter(myteam=result)
            
#             if not item_infos.exists():
#                 return JsonResponse({'message': 'Item not found'}, status=404)
            
#             items_data = []
            
#             for item_info in item_infos:
#                 stockin = Stockin.objects.filter(myteam=item_info)
#                 stockout = Stockout.objects.filter(myteam=item_info)
#                 adjust = Adjust.objects.filter(myteam=item_info)
                
#                 item_stockin_quantity = sum(item.initial_quantity for item in stockin)
#                 item_stockout_quantity = sum(item.initial_quantity for item in stockout)
#                 item_adjust_quantity = sum(item.initial_quantity for item in adjust)
                
#                 balance = item_stockin_quantity + item_stockout_quantity + item_adjust_quantity
                
#                 item_data = {
#                     "image": item_info.image.url if item_info.image else None,
#                     "name": item_info.Name,
#                     "Price": item_info.Price,
#                     "cost": item_info.cost,
#                     "initial_quantity": item_info.initial_quantity,
#                     "Stock_in": item_stockin_quantity,
#                     "stock_out": item_stockout_quantity,
#                     "adjust": item_adjust_quantity,
#                     "Balance": balance
#                 }
#                 items_data.append(item_data)

                
            
#             response_data = {
#                 "Item_Information": items_data 
#             }
                
#             return JsonResponse(response_data)
#         except Myteam.DoesNotExist:
#             return JsonResponse({'message': 'Team not found'}, status=404)
#         except Exception as e:
#             return JsonResponse({'message': 'error', 'details': str(e)}, status=500)
#     else:
#         return JsonResponse({'message': 'Invalid request method'}, status=405)

# from django.http import JsonResponse
# from django.views.decorators.csrf import csrf_exempt
# from .models import Myteam, Item_Information, Stockin, Stockout, Adjust  # Adjust the import based on your project structure

@csrf_exempt
def summary(request, id):
    if request.method == 'GET':
        try:
            result = Myteam.objects.get(id=id)
            item_infos = Item_Information.objects.filter(myteam=result)
            
            if not item_infos.exists():
                return JsonResponse({'message': 'Item not found'}, status=404)
            
            items_data = []
            
            for item_info in item_infos:
                stockin = Stockin.objects.filter(myteam=item_info)
                stockout = Stockout.objects.filter(myteam=item_info)
                adjust = Adjust.objects.filter(myteam=item_info)
                
                item_stockin_quantity = sum(item.initial_quantity for item in stockin)
                item_stockout_quantity = sum(item.initial_quantity for item in stockout)
                item_adjust_quantity = sum(item.initial_quantity for item in adjust)
                
                balance = item_stockin_quantity + item_stockout_quantity + item_adjust_quantity
                
                item_data = {
                    "image": item_info.image.url if item_info.image else None,
                    "name": item_info.Name,
                    "Price": item_info.Price,
                    "cost": item_info.cost,
                    "initial_quantity": item_info.initial_quantity,
                    "Stock_in": item_stockin_quantity,
                    "stock_out": item_stockout_quantity,
                    "adjust": item_adjust_quantity,
                    "Balance": balance
                }
                items_data.append(item_data)

            # Calculate totals for the entire team
            total_team_stockin_quantity = sum(item.initial_quantity for item in Stockin.objects.filter(myteam__in=item_infos))
            total_team_stockout_quantity = sum(item.initial_quantity for item in Stockout.objects.filter(myteam__in=item_infos))
            total_team_adjust_quantity = sum(item.initial_quantity for item in Adjust.objects.filter(myteam__in=item_infos))
            item_count = item_infos.count()
            
            team_data = {
                "total_team_stock_in": total_team_stockin_quantity,
                "total_team_stock_out": total_team_stockout_quantity,
                "total_team_adjust": total_team_adjust_quantity,
                "Item_count": item_count
            }
            
            response_data = {
                "Item_Information": items_data,
                "Team_Summary": team_data
            }
                
            return JsonResponse(response_data)
        except Myteam.DoesNotExist:
            return JsonResponse({'message': 'Team not found'}, status=404)
        except Exception as e:
            return JsonResponse({'message': 'error', 'details': str(e)}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)

    
######################################################## Dashboard #########################################################################################


# @csrf_exempt
# def uploadimage(request):
#     if request.method == 'POST':
#         try:
#             user_image = request.FILES.get('photo')
#             if not user_image:
#                 return JsonResponse({'message': 'No image file provided'})

#             new_image = image.objects.create(image=user_image)
#             new_image.save()
            
#             return JsonResponse({'message': 'Data saved successfully'})
#         except Exception as e:
#             return JsonResponse({'message': 'Data not saved', 'error': str(e)})
#     else:
#         return JsonResponse({'message': 'Invalid request method'})



############################################### Transaction #################################################################

@csrf_exempt
def transaction(request, id):
    if request.method == 'GET':
        try:
            start = Myteam.objects.get(id=id)
            
            value = Item_Information.objects.filter(myteam=start)
            for item in value:
                attribute = ItemAttribute.objects.filter(item=item)
                attribute_list = [{"attribute_value": i.attribute_value} for i in attribute]
                transactions = Transaction.objects.all()
                data = []

                myuser = Myuser.objects.all()
                for i in myuser:
                    for transaction in transactions:
                        stockins = Stockin.objects.filter(transaction=transaction, myteam__myteam=start)
                        stockouts = Stockout.objects.filter(transaction=transaction, myteam__myteam=start)
                        adjusts = Adjust.objects.filter(transaction=transaction, myteam__myteam=start)
                        
                        total_initial_quantity_stockin = sum(stockin.initial_quantity for stockin in stockins)
                        total_initial_quantity_stockout = sum(stockout.initial_quantity for stockout in stockouts)
                        total_initial_quantity_adjust = sum(adjust.initial_quantity for adjust in adjusts)
                        
                        transaction_entries = []

                        # Stock In Entries
                        for stock in stockins:
                            item_info = stock.myteam
                            transaction_entries.append({
                                "type": "Stock In",
                                "id": stock.id,
                                "supplier": stock.supplier,
                                "item_name": item_info.Name,
                                "cost": item_info.cost,
                                "price": item_info.Price,
                                "Barcode": item_info.Barcode,
                                "attribute": attribute_list,
                                "image": item_info.image.url if item_info.image else None,
                                "initial_quantity": stock.initial_quantity,
                                "created_at": stock.created_at,
                                "item_available": item_info.initial_quantity
                            })

                        # Stock Out Entries
                        for stock in stockouts:
                            item_info = stock.myteam
                            transaction_entries.append({
                                "type": "Stock Out",
                                "id": stock.id,
                                "customer": stock.customer,
                                "item_name": item_info.Name,
                                "cost": item_info.cost,
                                "price": item_info.Price,
                                "Barcode": item_info.Barcode,
                                "attribute": attribute_list,
                                "image": item_info.image.url if item_info.image else None,
                                "initial_quantity": stock.initial_quantity,
                                "created_at": stock.created_at,
                                "item_available": item_info.initial_quantity
                            })

                        # Adjust Entries
                        for adjust in adjusts:
                            item_info = adjust.myteam
                            transaction_entries.append({
                                "type": "Adjust",
                                "id": adjust.id,
                                "item_name": item_info.Name,
                                "cost": item_info.cost,
                                "price": item_info.Price,
                                "Barcode": item_info.Barcode,
                                "attribute": attribute_list,
                                "image": item_info.image.url if item_info.image else None,
                                "initial_quantity": adjust.initial_quantity,
                                "created_at": adjust.created_at,
                                "item_available": item_info.initial_quantity
                            })

                        if transaction_entries:
                            entry_type = transaction_entries[0]['type'].replace(" ", "")
                            entry_data = {
                                "transaction_id": str(transaction.transaction_id),
                                "nick_name": i.nick_name,
                                "item_length": len(transaction_entries),
                                "created_at": transaction.created_at,
                                "type": entry_type,
                                "entries": transaction_entries
                            }
                            
                            if entry_type == "StockIn":
                                entry_data["total_initial_quantity"] = total_initial_quantity_stockin
                            elif entry_type == "StockOut":
                                entry_data["total_initial_quantity"] = total_initial_quantity_stockout
                            elif entry_type == "Adjust":
                                entry_data["total_initial_quantity"] = total_initial_quantity_adjust
                            
                            data.append(entry_data)

                    return JsonResponse({"data": data}, safe=False)
        
        except Myteam.DoesNotExist:
            return JsonResponse({'message': 'Myteam not found'}, status=404)
        except Exception as e:
            return JsonResponse({'message': 'Error retrieving transactions', 'error': str(e)}, status=500)
    
    else:
        return JsonResponse({'message': 'Method not allowed'}, status=405)



    
from django.utils import timezone 
@csrf_exempt
def today_transaction(request,id):
    if request.method == 'GET':
        try:
            start = Myteam.objects.get(id=id)
            value = Item_Information.objects.filter(myteam=start)
            print(filter)
            data = []
            today = timezone.now().date()

            for item in value:
                entries = []

                stockins = Stockin.objects.filter(myteam=item, created_at__date=today).order_by('-created_at')
                users = Myuser.objects.all()
                for user in users:
                    for stock in stockins:
                        entries.append({
                            "type": "Stock In",
                            "supplier": stock.supplier,
                            "item_name": item.Name,
                            "initial_quantity": stock.initial_quantity,
                            "created_at": stock.created_at,
                            "user": user.nick_name,
                            "item_available": item.initial_quantity
                        })

                stockouts = Stockout.objects.filter(myteam=item, created_at__date=today).order_by('-created_at')
                for user in users:
                    for stock in stockouts:
                        entries.append({
                            "type": "Stock Out",
                            "customer": stock.customer,
                            "item_name": item.Name,
                            "initial_quantity": stock.initial_quantity,
                            "created_at": stock.created_at,
                            "user": user.nick_name,
                            "item_available": item.initial_quantity
                        })

                adjusts = Adjust.objects.filter(myteam=item, created_at__date=today).order_by('-created_at')
                for user in users:
                    for adjust in adjusts:
                        entries.append({
                            "type": "Adjust",
                            "item_name": item.Name,
                            "initial_quantity": adjust.initial_quantity,
                            "created_at": adjust.created_at,
                            "user": user.nick_name,
                            "item_available": item.initial_quantity
                        })

                # if entries:
                #     # Sort all entries by created_at in descending order
                #     entries.sort(key=lambda x: x['created_at'], reverse=True)
                #     data.extend(entries)  # Flattening the list

            return JsonResponse({"data": entries}, safe=False)
        except Myteam.DoesNotExist:
            return JsonResponse({'message': 'Myteam not found'}, status=404)
        except Exception as e:
            return JsonResponse({'message': 'There is no stockin, stockout and adjust in this team id.'}, status=500)
    else:
        return JsonResponse({'message': 'Method not allowed'}, status=405)
        
################################################## User ################################################################################

@csrf_exempt
def user_show(request, Email):
    if request.method == 'GET':
        try:
            user = Myuser.objects.get(email=Email)
           
            value = {
                "name":user.nick_name,
                "Email":user.email,
                "image":user.image.url
                
            }
            # data.append(value)
        
            return JsonResponse(value, safe=False)
    
        
        except Myuser.DoesNotExist:
            return JsonResponse({'message': 'User not found'}, status=404)
        except Exception as e:
            return JsonResponse({'message': 'Some error occurred', 'error': str(e)}, status=500)
    
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)

        
@csrf_exempt
def user_edit(request, Email):
    if request.method == 'POST':
        name = request.POST.get('name')
        image = request.FILES.get('image')
        
        try:
            result = Myuser.objects.get(email=Email)
            if result:
                result.nick_name = name
                if image:
                    result.image = image
                result.save()
                return JsonResponse({'message': 'Data updated successfully.'})
            else:
                return JsonResponse({'message': 'User not found.'}, status=404)
        except Myuser.DoesNotExist:
            return JsonResponse({'message': 'User not found.'}, status=404)
        except Exception as e:
            return JsonResponse({'message': 'Error', 'error': str(e)}, status=500)
    else:
        return JsonResponse({'message': 'Method not allowed'}, status=405)


@csrf_exempt
def team_delete(request, id):
    if request.method == 'DELETE':
        try:
            result = Myteam.objects.filter(id=id)
            if result.exists():
                result.delete()
                return JsonResponse({'message': 'Delete successful'})
            else:
                return JsonResponse({'message': 'ItemAttribute not found'}, status=404)
        except Exception as e:
            return JsonResponse({'message': str(e)}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)

from django.contrib.contenttypes.models import ContentType

############################################# Permission ############################################################
    
@csrf_exempt
def assign_permission(request, id):
    if request.method == 'POST':
        email = request.POST.get('email')
        image = request.FILES.get('image')
        permission = request.POST.get('permission').lower()
        
        print(email, image, permission)
        
        if permission not in dict(Permiteduser.PERMISSION_TYPES):
            additional_permission = permission
            permission = None
        else:
            additional_permission = None
        
        try:
            send_permission_mail(email)
            print("email send successfully")
            
            team = Myteam.objects.get(id=id)
            if team:
                user = Permiteduser.objects.create(
                    email=email,
                    image=image,
                    permission=permission,
                    additional_permission=additional_permission,
                    myteam=team
                )
                user.save()

                return JsonResponse({'message': 'Permission added successfully'})
        except Myteam.DoesNotExist:
            return JsonResponse({'message': 'Team not found'}, status=404)
        except Exception as e:
            return JsonResponse({'message': f'An error occurred: {e}'}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)
    
@csrf_exempt
def resend_email(request,id):
    if request.method == 'GET':
        
        try:
            user = Permiteduser.objects.get(id=id)
            email = user.email
            
            send_permission_mail(email)
            print("email send successfully")
            return JsonResponse({'message': 'Email resent successfully'}, status=200)
        
        except Permiteduser.DoesNotExist:
            return JsonResponse({'message': 'Permiteduser not found'}, status=404)
        except Exception as e:
            return JsonResponse({'message': f'An error occurred: {e}'}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405) 
    
@csrf_exempt
def delete_permiteduser(request, id):
    if request.method == 'DELETE':
        try:
            result = Permiteduser.objects.filter(id=id)
            if result.exists():
                result.delete()
                return JsonResponse({'message': 'Delete successful'})
            else:
                return JsonResponse({'message': 'ItemAttribute not found'}, status=404)
        except Exception as e:
            return JsonResponse({'message': str(e)}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)


@csrf_exempt
def custom_permission(request, id):                                      #(role 1st)
    if request.method == 'POST':
        jsondata = JSONParser().parse(request)
        name = jsondata.get('name')
        permissions = jsondata.get('permissions')
        print(name, permissions)
        
        try:
            myteam = Myteam.objects.get(id=id)
            
            role = Role.objects.create(name=name, permissions=permissions, role=myteam)
            
            return JsonResponse({'message': 'Data saved successfully.'}, status=201)

        except Myteam.DoesNotExist:
            return JsonResponse({'message': 'Myteam not found'}, status=404)
        except Exception as e:
            return JsonResponse({'message': 'error', 'details': str(e)}, status=400)
    
    else:
        return JsonResponse({'error': 'Invalid request method.'}, status=405)


@csrf_exempt
def custom_permission_view(request,id):                      #(role)
    if request.method == 'GET':
        try:
            team = Myteam.objects.get(id=id)
            if team:
                data = Role.objects.all()
                mm = []
                for group in data:
                    view = {
                        "name": group.name,
                        "member":"No Members"
                    }
                    mm.append(view)
                return JsonResponse(mm, safe=False)
        except Exception as e:
            return JsonResponse({'message': 'error', 'details': str(e)}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)
    
    
from django.core import signing

@csrf_exempt
def change_status(request, token):
    try:
        data = signing.loads(token)
        model_id = data['model_id']

        model_instance = Permiteduser.objects.get(email=model_id)
        model_instance.permission_status = 'yes'
        model_instance.save()

        return JsonResponse({'message': 'Status changed successfully'})
    except signing.BadSignature:
        return JsonResponse({'message': 'Invalid token'}, status=400)
    except Permiteduser.DoesNotExist:
        return JsonResponse({'message': 'Model instance not found'}, status=404)
    except Exception as e:
        return JsonResponse({'message': f'An error occurred: {e}'}, status=500)
    
    
groups_permissions = {
    "Admin":"Admin",
    "Viewer":"Viewer"
}

def get_groups_permissions(request):
    if request.method == 'GET':
        roles = Role.objects.all()
        role_names = [role.name for role in roles]
        group_keys = list(groups_permissions.keys())
        combined_list = group_keys + role_names
        # data = {
        #     "roles":role_names,
        #     "groups":group_keys
        # }
        
        # return JsonResponse(data ,safe=False)
        return JsonResponse({'groups': combined_list}, status=200)
        # return JsonResponse({'groups': group_keys, 'roles': role_names}, status=200)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)
    

@csrf_exempt
def view_custom_permission(request,id):                         #(role)
    if request.method == 'GET':
        try:
            data = Permiteduser.objects.get(id=id)
            if data:
                role = Role.objects.filter(role=data)
                vv = []
                for i in role:
                    xx = {
                        "email" : data.email,
                        "name": i.name   
                    }
                    vv.append(xx)
            return JsonResponse(vv, safe = False)
        except:
            return JsonResponse({'message':'error'})
    
# @csrf_exempt
# def edit_custom_permission(request):
#     if request.method == 'POST':
#         jsondata = JSONParser().parse(request)
#         permission = 
        
 
@csrf_exempt
def invited_user_list(request, id):
    if request.method == 'GET':
        try:
            team = Myteam.objects.get(id=id)
            if team:
                data = Permiteduser.objects.filter(myteam=team)
                val = []

                for i in data:
                    response = {
                        "id": i.id,
                        "email": i.email,
                        "permissions": i.permission,
                        "permission_status": i.permission_status
                    }
                    val.append(response)

                return JsonResponse(val, safe=False)
        except Myteam.DoesNotExist:
            return JsonResponse({'message': 'Team id not found'}, status=401)
        except Exception as e:
            return JsonResponse({'message': f'An error occurred: {e}'}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)
    
    
    
#############################################################################################################################

import cv2
import numpy as np

@csrf_exempt
def find_similar_image(request):
    if request.method == 'POST':
        image_file = request.FILES.get('image')
        if not image_file:
            return JsonResponse({'message': 'Image is required'}, status=400)

        # Read the uploaded image
        file_bytes = np.frombuffer(image_file.read(), np.uint8)
        uploaded_image = cv2.imdecode(file_bytes, cv2.IMREAD_COLOR)
        gray = cv2.cvtColor(uploaded_image, cv2.COLOR_BGR2GRAY)
        sift = cv2.SIFT_create()
        _, uploaded_des = sift.detectAndCompute(gray, None)

        if uploaded_des is None:
            return JsonResponse({'message': 'No features found in the uploaded image'}, status=400)

        # Compare with stored images
        matches_list = []
        # threshold = 0.7 # Its for exact same
        threshold = 0.77  # Its for overal same

        for item in Item_Information.objects.all():
            if item.feature_vector:
                stored_des = np.frombuffer(item.feature_vector, dtype=np.float32).reshape(-1, 128)
                bf = cv2.BFMatcher()
                matches = bf.knnMatch(uploaded_des, stored_des, k=2)

                # Apply ratio test
                good_matches = [m for m, n in matches if m.distance < threshold * n.distance]

                if len(good_matches) > 10:  # Only consider matches with a significant number of good matches
                    score = sum([match.distance for match in good_matches]) / len(good_matches)
                    matches_list.append({'image_url': item.image.url, 'score': score})

        if matches_list:
            matches_list.sort(key=lambda x: x['score'])  # Sort by score (lower is better)
            return JsonResponse({'message': 'Images found', 'images': matches_list})
        else:
            return JsonResponse({'message': 'No similar image found'}, status=404)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)
    
############################################## Dashboard ########################################################################

from django.db.models import Sum

@csrf_exempt
def dashboard(request, id):
    if request.method == 'GET':
        try:
            team = Myteam.objects.get(id=id)
            items = Item_Information.objects.filter(myteam=team)

            total_initial_quantity = 0
            total_stockin_quantity = 0
            total_stockout_quantity = 0

            item_data = {}
            stockin_data = {}
            stockout_data = {}

            end_date = datetime.now()
            start_date = end_date - timedelta(days=30)
            
            formatted_start_date = start_date.strftime('%b %d, %Y')
            formatted_end_date = end_date.strftime('%b %d, %Y')
            date_range = f"{formatted_start_date} ~ {formatted_end_date}"

            for item in items:
                
                total_initial_quantity += item.initial_quantity

                stockin_records = Stockin.objects.filter(myteam=item)
                stockin_sum = stockin_records.aggregate(Sum('initial_quantity'))['initial_quantity__sum'] or 0
                total_stockin_quantity += stockin_sum

                stockout_records = Stockout.objects.filter(myteam=item)
                stockout_sum = stockout_records.aggregate(Sum('initial_quantity'))['initial_quantity__sum'] or 0 
                total_stockout_quantity += stockout_sum

                recent_stockin_records = stockin_records.filter(created_at__range=[start_date, end_date])
                recent_stockout_records = stockout_records.filter(created_at__range=[start_date, end_date])

                item_created_date = item.created_at.strftime('%b %d')
                item_data[item_created_date] = item_data.get(item_created_date, 0) + item.initial_quantity

                for record in recent_stockin_records:
                    stockin_created_date = record.created_at.strftime('%b %d')
                    stockin_data[stockin_created_date] = stockin_data.get(stockin_created_date, 0) + record.initial_quantity

                for record in recent_stockout_records:
                    stockout_created_date = record.created_at.strftime('%b %d')
                    stockout_data[stockout_created_date] = stockout_data.get(stockout_created_date, 0) + record.initial_quantity

            response_data = {
                "Item_data": list(item_data.values()),
                "Item_created_at": list(item_data.keys()),
                "Stockin_data": list(stockin_data.values()),
                "Stockin_created_at": list(stockin_data.keys()),
                "stockout_data": list(stockout_data.values()),
                "stockout_created_at": list(stockout_data.keys()),
                "date_range": date_range,
                "data": {
                    "total_initial_quantity": total_initial_quantity,
                    "total_stockin": total_stockin_quantity,
                    "total_stockout": total_stockout_quantity
                }
            }

            return JsonResponse(response_data, safe=False)

        except Myteam.DoesNotExist:
            return JsonResponse({'message': 'Team not found'}, status=404)
        except Exception as e:
            return JsonResponse({'message': 'Error', 'details': str(e)}, status=500)
    else:
        return JsonResponse({'message': 'Invalid request method'}, status=405)

    
    
@csrf_exempt
def yesterday_data(request, id):
    if request.method == 'GET':
        
        try:
            team = Myteam.objects.get(id=id)
            items = Item_Information.objects.filter(myteam=team)
            
            end_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            start_date = end_date - timedelta(days=1)
            
            stockin = Stockin.objects.filter(myteam__in=items, created_at__range=[start_date, end_date])
            stockout = Stockout.objects.filter(myteam__in=items, created_at__range=[start_date, end_date])
            
            total_inventory = sum(i.initial_quantity for i in items)
            total_stockin = sum(i.initial_quantity for i in stockin)
            total_stockout = sum(i.initial_quantity for i in stockout)
            
            data = {
                "total_yesterday_inventory": total_inventory,
                "total_yesterday_stockin": total_stockin,
                "total_yesterday_stockout": total_stockout
            }
            
            return JsonResponse(data, safe=False)
        except Exception as e:
            return JsonResponse({'message': 'error', 'details': str(e)}, status=500)



#################################################################################################################################
from rest_framework import viewsets, permissions, status
from rest_framework.response import Response
from django.contrib.auth.models import Group, Permission
from django.contrib.auth import get_user_model

User = get_user_model()

class userViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    permission_classes = [permissions.IsAuthenticated]
    # Define your serializer here

    def add_permission(self, request, pk=None):
        user = self.get_object()
        permission = request.data.get('permission')
        if permission:
            try:
                perm = Permission.objects.get(codename=permission)
                user.user_permissions.add(perm)
                return Response({'status': 'permission added'}, status=status.HTTP_200_OK)
            except Permission.DoesNotExist:
                return Response({'status': 'permission not found'}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'status': 'permission not provided'}, status=status.HTTP_400_BAD_REQUEST)
    def remove_permission(self, request, pk=None):
        user = self.get_object()
        permission = request.data.get('permission')
        if permission:
            perm = Permission.objects.get(codename=permission)
            user.user_permissions.remove(perm)
            return Response({'status': 'permission removed'}, status=status.HTTP_200_OK)
        return Response({'status': 'permission not provided'}, status=status.HTTP_400_BAD_REQUEST)

from .serializers import StockinSerializer 

class StockinViewSet(viewsets.ModelViewSet):
    queryset = Stockin.objects.all()
    serializer_class = StockinSerializer
    permission_classes = [permissions.DjangoModelPermissions]

    def create(self, request, *args, **kwargs):
        if not request.user.has_perm('your_app.can_stockin'):
            return Response({'detail': 'You do not have permission to perform this action.'}, status=status.HTTP_403_FORBIDDEN)
        return super().create(request, *args, **kwargs)
    
    
    